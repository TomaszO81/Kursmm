import openpyxl

plik = openpyxl.load_workbook('example.xlsx')
print(type(plik))

# lista arkuszy
arkusze = plik.get_sheet_names()
print(arkusze)

#konkretny arkusz

arkusz = plik.get_sheet_by_name('Owoce')
print("nazawa arkusza", arkusz.title)

#aktywny arkusz
print('Aktywny arkusz:', plik.active)

#komorki
komorka = arkusz['A1']
print(komorka)
print(komorka.value)

# koordynaty komorki
print('Adres komorki:', komorka.coordinate)
print('kolumna komorki', komorka.column)
print('wiersz komorki', komorka.row)

#konkretna komorka arkusza
print(arkusz.cell(column=2, row=1))

#rozmiar arkusza
print('Ostania kolumna:', arkusz.max_column)


#zamiana litery na cyfry
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(20))

#zakresy danych
print(arkusz['A1':'C3'])
for wiersz in arkusz['A1':'C3']:
    for kom in wiersz:
        print('Komorka {} ma wartość {}'.format(kom.coordinate, kom.value))
        print('-------------koniec wiersza-----------------')







plik.close()